"""Import Excel central sous /api/partners/{partner_id}/imports."""
import os

from fastapi import APIRouter, Depends, UploadFile, File, Form, Query
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.errors import NotFoundError
from app.api.deps import get_current_user, get_partner_context, require_roles
from app.crud.import_batch_crud import import_batch_crud
from app.models.user import User
from app.security.permissions import Role, IMPORT_ROLES
from app.schemas.import_batch import ImportBatchOut, ImportValidationResult
from app.schemas.pagination import Page
from app.services.import_validation_service import validate_import, apply_import, REQUIRED_COLUMNS
from app.services.odi_import_service import import_zone_file, import_stock_file

router = APIRouter(prefix="/api/partners/{partner_id}/imports", tags=["Import Excel"])


@router.post("/validate", response_model=ImportValidationResult)
async def validate_import_route(
    entity_type: str = Form(...),
    file: UploadFile = File(...),
    partner_id: int = Depends(get_partner_context),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*IMPORT_ROLES)),
):
    content = await file.read()
    return validate_import(
        db, partner_id=partner_id, user_id=user.id, entity_type=entity_type,
        filename=file.filename, file_bytes=content,
    )


@router.post("/{batch_id}/apply")
def apply_import_route(batch_id: int, partner_id: int = Depends(get_partner_context),
                        db: Session = Depends(get_db),
                        user: User = Depends(require_roles(*IMPORT_ROLES))):
    result = apply_import(db, partner_id=partner_id, user_id=user.id, batch_id=batch_id)
    return {"batch": ImportBatchOut.model_validate(result["batch"]), "applied_rows": result["applied_rows"]}


@router.get("/{batch_id}", response_model=ImportBatchOut)
def get_batch(batch_id: int, partner_id: int = Depends(get_partner_context),
              db: Session = Depends(get_db), _user: User = Depends(get_current_user)):
    from app.core.errors import NotFoundError
    batch = import_batch_crud.get(db, batch_id)
    if not batch or batch.partner_id != partner_id:
        raise NotFoundError("Lot d'import introuvable dans ce Partenaire.")
    return batch


@router.get("", response_model=Page[ImportBatchOut])
def list_batches(partner_id: int = Depends(get_partner_context),
                  skip: int = 0, limit: int = Query(default=100, le=500),
                  db: Session = Depends(get_db),
                  _user: User = Depends(get_current_user)):
    return import_batch_crud.list_paginated(db, skip=skip, limit=limit, partner_id=partner_id)


# Exemple-type des gabarits Excel téléchargeables.
# Reflet du référentiel réel (Master Color = PART-MC), Camtel Express (PART-001) ayant été retiré.
SAMPLE_ROWS = {
    "PARTNER": {"code_partenaire": "PART-MC", "name": "Master Color"},
    "DSM": {"matricule": "DSM-DLA-01", "full_name": "Jean Marc"},
    "POS": {"code_pos": "POS-0001", "name": "Kiosque Akwa", "dsm_matricule": "DSM-DLA-01",
             "date_creation": "2026-01-15", "date_expiration": "2026-12-31"},
    "BTS": {"code_bts": "BTS-DLA-01", "operateur": "CAMTEL", "technologie": "4G",
             "capacite_max": 1000, "latitude": 4.0511, "longitude": 9.7679},
    "BTS_RELEVE": {"bts_code": "BTS-DLA-01", "charge": 62.5, "taux_saturation": 48.0,
                    "rendement": 91.2, "date_releve": "2026-08-01"},
    "SIM": {"iccid": "89237010000000000001", "pos_code": "POS-0001"},
    "PRIME_PERIOD": {"code": "P2026-T3", "label": "Trimestre 3", "start_date": "2026-07-01",
                      "end_date": "2026-09-30"},
    "PRIME": {"pos_code": "POS-0001", "prime_period_code": "P2026-T3", "montant": 15000},
    "REQUETE": {"external_id": "EXT-REQ-001", "type_requete": "AJOUT", "titre": "Ajout POS Akwa",
                 "description": "Besoin de 2 POS supplémentaires", "priorite": "NORMALE",
                 "entite_en_charge": "AC Bépanda"},
    "SALES_TARGET": {"month": "2026-08-01", "creation_target": 40, "redeployment_target": 10,
                     "sell_out_target": 60, "loading_target": 50,
                     "creation_stock_initial": 10, "redeployment_stock_initial": 5},
}


@router.get("/templates/{entity_type}")
def download_template(entity_type: str,
                      _partner_id: int = Depends(get_partner_context),
                      _user: User = Depends(get_current_user)):
    """Télécharge le gabarit Excel officiel d'un type d'entité.

    Colonnes = REQUIRED_COLUMNS du service de validation (+ colonnes
    optionnelles documentées pour certains types), avec une ligne
    d'exemple pour guider la saisie terrain.
    """
    from io import BytesIO

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook

    key = entity_type.strip().upper()
    required = REQUIRED_COLUMNS.get(key)
    if not required:
        raise NotFoundError(
            f"Type d'entite '{entity_type}' non supporte. "
            f"Valeurs possibles : {list(REQUIRED_COLUMNS.keys())}"
        )

    optional = {
        "DSM": ["zone"],
        "POS": ["address", "zone", "stock_initial"],
        "BTS_RELEVE": ["debit", "connexions", "latence"],
        "SIM": ["status"],
        "PRIME_PERIOD": ["statut"],
        "PRIME": [],
        "REQUETE": ["nombre_demande"],
        "SALES_TARGET": ["creation_target", "redeployment_target", "sell_out_target", "loading_target",
                          "creation_stock_initial", "redeployment_stock_initial"],
    }.get(key, [])

    columns = required + [c for c in optional if c not in required]
    sample = SAMPLE_ROWS.get(key, {})

    wb = Workbook()
    ws = wb.active
    ws.title = key
    ws.append(columns)
    ws.append([sample.get(c, "") for c in columns])

    # Style léger : entête en gras + largeurs ajustées.
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for col_idx, col in enumerate(columns, start=1):
        width = max(len(col), len(str(sample.get(col, "")))) + 4
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(width, 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="gabarit_{key.lower()}.xlsx"'},
    )


@router.get("/{batch_id}/report")
def download_error_report(batch_id: int, partner_id: int = Depends(get_partner_context),
                           db: Session = Depends(get_db), _user: User = Depends(get_current_user)):
    """Telecharge le rapport d'erreurs (JSON) d'un lot d'import donne."""
    batch = import_batch_crud.get(db, batch_id)
    if not batch or batch.partner_id != partner_id:
        raise NotFoundError("Lot d'import introuvable dans ce Partenaire.")
    if not batch.error_report_path or not os.path.exists(batch.error_report_path):
        raise NotFoundError("Rapport d'erreurs indisponible pour ce lot.")
    return FileResponse(
        batch.error_report_path,
        media_type="application/json",
        filename=f"rapport_erreurs_lot_{batch.id}.json",
    )


# --- Import de donnees reels (format ZONE / STOCK) ---

@router.post("/zone")
async def import_zone_route(
    file: UploadFile = File(...),
    partner_id: int = Depends(get_partner_context),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*IMPORT_ROLES)),
):
    """Importe un fichier ZONE Excel (format geographique BTS avec bornes N/E/S/W).

    Format attendu : colonnes SN, PARTNERS, BTS CODE NAME, PROMINENT SITES,
    BOUNDARIES, QUARTER, STREET, RADIUS, GPS COORDINATES, COVERAGE,
    CAPACITY, TRAFFIC VOLUME, POS.

    Chaque BTS est representee par 4 lignes de bornes (N, E, S, W) +
    une ligne principale avec les donnees GPS et couverture.
    """
    from app.services import audit_service

    content = await file.read()
    result = import_zone_file(db, partner_id=partner_id, file_bytes=content)

    audit_service.log_action(
        db, user_id=user.id, partner_id=partner_id, action="IMPORT_ZONE",
        entity_type="BTS", entity_id=None,
        details=f"Import ZONE : {result['created']} BTS cree(s), {result['updated']} mis a jour",
    )

    return result


@router.post("/stock")
async def import_stock_route(
    file: UploadFile = File(...),
    partner_id: int = Depends(get_partner_context),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(*IMPORT_ROLES)),
):
    """Importe un fichier STOCK Excel (format hierarchique DSM->POS avec solde SIM).

    Format attendu : colonnes Level, Organization Name, code (col3),
    parent/type (col4), zone_type, ..., Parent Organization Name,
    Current Balance.

    Niveaux :
    - 4 : Partenaire/Zone maitre
    - 5 : DSM (org_id, matricule, zone_type, solde)
    - 6 : POS (org_id, code, parent DSM, zone_type, solde)
    """
    from app.services import audit_service

    content = await file.read()
    result = import_stock_file(db, partner_id=partner_id, file_bytes=content)

    audit_service.log_action(
        db, user_id=user.id, partner_id=partner_id, action="IMPORT_STOCK",
        entity_type="DSM", entity_id=None,
        details=(f"Import STOCK : {result['created_dsm']} DSM cree(s), "
                 f"{result['created_pos']} POS cree(s)"),
    )

    return result
